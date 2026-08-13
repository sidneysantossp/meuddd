from collections import Counter
from pathlib import Path
from urllib.parse import urlparse
from openpyxl import load_workbook
import json

source = Path("/home/ubuntu/upload/www.meuddd.com.br_internal_images_are_broken_20260813.xlsx")
workbook = load_workbook(source, read_only=True, data_only=True)
sheet = workbook.active

headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
rows = [dict(zip(headers, values)) for values in sheet.iter_rows(min_row=2, values_only=True)]

image_urls = Counter(row["Image URL"] for row in rows)
http_codes = Counter(str(row["HTTP Code"]) for row in rows)
page_urls = Counter(row["Page URL"] for row in rows)
extensions = Counter(Path(urlparse(url).path).suffix.lower() or "(sem extensão)" for url in image_urls)
path_groups = Counter()
for page in page_urls:
    path = urlparse(page).path
    if path.startswith("/cidade/"):
        path_groups["municípios"] += 1
    elif path.startswith("/estado/"):
        path_groups["estados"] += 1
    elif path.startswith("/ddd/"):
        path_groups["DDDs"] += 1
    elif path.startswith("/guia/"):
        path_groups["guias"] += 1
    elif path == "/":
        path_groups["homepage"] += 1
    else:
        path_groups["outras"] += 1

summary = {
    "linhas_de_imagem_com_falha": len(rows),
    "imagens_unicas": len(image_urls),
    "paginas_afetadas_unicas": len(page_urls),
    "codigos_http": http_codes,
    "imagens_por_ocorrencia": image_urls,
    "extensoes": extensions,
    "paginas_por_grupo": path_groups,
    "amostra_de_paginas": list(page_urls.keys())[:5],
}
print(json.dumps(summary, ensure_ascii=False, indent=2))
